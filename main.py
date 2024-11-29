import pandas as pd
import shutil  # Para copiar el archivo
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Cargar los archivos Excel
#ORIGEN
archivo_data = "Book1.xlsx"
hoja_data = "Sheet1"

#DESTINO
archivo_actualizar = "Book2.xlsx"
hoja_actualizar = "nuevosdatos"

# Crear un respaldo del archivo B antes de modificarlo
respaldo_b = datetime.now().strftime('%Y%m%d-%H-%M%S-')+"Respaldo"+archivo_actualizar
shutil.copy(archivo_actualizar, respaldo_b)
print(f"Respaldo creado: {respaldo_b}")

# Solicitar al usuario las hojas a usar
#hoja_a = input(f"Ingresa el nombre de la hoja del archivo A ({archivo_data}): ")
#hoja_b = input(f"Ingresa el nombre de la hoja del archivo B ({archivo_actualizar}): ")


# Leer las hojas especificadas
df_datos = pd.read_excel(archivo_data, sheet_name=hoja_data)
df_actualizar = pd.read_excel(archivo_actualizar, sheet_name=hoja_actualizar)

# Diccionario de mapeo entre columnas de A y B
# 'columna_data': 'columna_actualizar'
#ORIGEN
refcolumDatos="nombre"
#DESTINO
refcolumActualizar="producto"

#DATOS DE A ===> B
mapeo_columnas = {
    "asesor": "asesor",
    "precio": "precio",
    "caramba": "gorrion",
}

# Iterar por el mapeo y actualizar las columnas correspondientes
for columna_data, columna_actualizar in mapeo_columnas.items():
    # Asegurar que la columna existe en ambos DataFrames
    if columna_data in df_datos.columns and columna_actualizar in df_actualizar.columns:
        # Actualizar df_actualizar con valores de df_datos cuando coincida la columna 'NOMBRE'
        df_actualizar.loc[df_actualizar[refcolumActualizar].isin(df_datos[refcolumDatos]), columna_actualizar] = \
            df_actualizar[refcolumActualizar].map(df_datos.set_index(refcolumDatos)[columna_data])

# Sobrescribir el archivo original de B con los cambios
with pd.ExcelWriter(archivo_actualizar, mode="a", if_sheet_exists="replace") as writer:
    df_actualizar.to_excel(writer, sheet_name=hoja_actualizar, index=False)

print(f"El archivo {archivo_actualizar} ha sido modificado directamente.")

# Identificar filas en origen sin coincidencias con destino
no_coinciden = ~df_datos[refcolumDatos].isin(df_actualizar[refcolumActualizar])

# Aplicar fondo rojo a las filas sin coincidencias en archivo_data usando openpyxl
wb = load_workbook(archivo_data)
ws = wb[hoja_data]

# Crear un estilo de fondo rojo
fondo_rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Iterar sobre las filas en el DataFrame y resaltar en Excel las que no tienen coincidencia
for index, no_coincidencia in enumerate(no_coinciden, start=2):  # start=2 para saltar encabezado
    if no_coincidencia:  # Si no hay coincidencia
        for col in range(1, ws.max_column + 1):  # Aplicar el fondo a todas las celdas de la fila
            ws.cell(row=index, column=col).fill = fondo_rojo

# Guardar el archivo Excel con las celdas marcadas
wb.save(archivo_data)

print(f"El archivo {archivo_data} ha sido modificado directamente. Las filas sin coincidencia tienen fondo rojo.")